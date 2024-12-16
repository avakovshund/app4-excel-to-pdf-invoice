import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

# Get an Excel filenames
filepaths = glob.glob("Excel Invoices/*.xlsx")

for filepath in filepaths:

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(False, margin=0)

    filename = Path(filepath).stem
    number, date = filename.split("-")

    # Add an Invoice number and Date from Excel filenames
    pdf.add_page()
    pdf.set_font(family="Times", style="B", size=16)

    pdf.cell(w=0, h=8, txt=f"Invoice nr. {number}", ln=1)
    pdf.cell(w=0, h=8, txt=f"Date {date}", ln=1)

    pdf.ln(8)

    # Add a headers of table
    excel_invoices = pd.read_excel(filepath, sheet_name="Sheet 1")
    columns = list(excel_invoices.columns)
    titles = [item.replace("_", " ").title() for item in columns]
    pdf.set_font(family="Times", style="B", size=12)
    pdf.cell(w=30, h=8, txt=titles[0], border=1)
    pdf.cell(w=45, h=8, txt=titles[1], border=1)
    pdf.cell(w=45, h=8, txt=titles[2], border=1)
    pdf.cell(w=30, h=8, txt=titles[3], border=1)
    pdf.cell(w=30, h=8, txt=titles[4], border=1, ln=1)

    # Add rows to the table
    for index, row in excel_invoices.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=45, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=45, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    # Add a total price cell
    pdf.set_font(family="Times", size=10)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=45, h=8, txt="", border=1)
    pdf.cell(w=45, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=f"{excel_invoices['total_price'].sum()}", border=1, ln=1)

    pdf.set_font(family="Times", style="B", size=12)
    pdf.cell(w=30, h=10, txt=f"The total price sum is {excel_invoices['total_price'].sum()}", ln=1)

    pdf.ln(8)

    pdf.set_font(family="Times", style="B", size=14)
    pdf.cell(w=37, h=8, txt=f"Avakov and Co.")
    pdf.image("avakov.jpg", w=8, h=8)

    pdf.output(f"PDF Invoices/{number}-{date}.pdf")

